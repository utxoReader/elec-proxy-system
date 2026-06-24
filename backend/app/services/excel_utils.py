"""
Shared Excel utility for import/export operations.

Provides common formatting, header styles, and the StreamingResponse
pattern used by all export endpoints.
"""

from io import BytesIO
from typing import Optional

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Style constants ──────────────────────────────────────────────────────

HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4F6AF5", end_color="4F6AF5", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

CELL_FONT = Font(name="微软雅黑", size=10)
CELL_ALIGNMENT = Alignment(horizontal="center", vertical="center")
CELL_ALIGNMENT_LEFT = Alignment(horizontal="left", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

# ── Helpers ──────────────────────────────────────────────────────────────


def create_workbook(sheet_title: str = "Sheet1") -> Workbook:
    """Create a new workbook with a single sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    return wb


def write_header(ws, headers: list[str], row: int = 1):
    """Write styled header row."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def write_row(ws, values: list, row: int, left_align_cols: Optional[set[int]] = None):
    """Write a data row with basic styling."""
    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font = CELL_FONT
        cell.border = THIN_BORDER
        if left_align_cols and col_idx in left_align_cols:
            cell.alignment = CELL_ALIGNMENT_LEFT
        else:
            cell.alignment = CELL_ALIGNMENT


def auto_width(ws, max_width: int = 40, min_width: int = 10):
    """Auto-adjust column widths."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = min_width
        for cell in col:
            if cell.value:
                length = len(str(cell.value))
                # CJK characters are roughly 2x width
                cjk_count = sum(1 for c in str(cell.value) if '\u4e00' <= c <= '\u9fff')
                length += cjk_count
                max_len = max(max_len, length)
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def save_to_response(wb: Workbook, filename: str) -> StreamingResponse:
    """Save workbook to a StreamingResponse for download."""
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def export_to_response(
    headers: list[str],
    rows: list[list],
    filename: str,
    sheet_title: str = "数据",
    left_align_cols: Optional[set[int]] = None,
) -> StreamingResponse:
    """Convenience: create workbook, write header + rows, return response."""
    wb = create_workbook(sheet_title)
    ws = wb.active
    write_header(ws, headers)
    for row_idx, values in enumerate(rows, 2):
        write_row(ws, values, row_idx, left_align_cols)
    auto_width(ws)
    return save_to_response(wb, filename)
