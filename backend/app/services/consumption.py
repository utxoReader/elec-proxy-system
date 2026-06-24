"""Consumption data business logic."""

from datetime import date, datetime, timezone
from decimal import Decimal
from typing import Optional

from fastapi import HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.crud.base import get_or_404, paginate_query
from app.models.consumption import (
    CustomerDailyConsumption,
    CustomerHourlyConsumption,
    Point96Data,
)
from app.schemas.consumption import (
    CustomerDailyConsumptionCreate,
    CustomerDailyConsumptionUpdate,
    CustomerHourlyConsumptionCreate,
    CustomerHourlyConsumptionUpdate,
    Point96DataCreate,
)
from app.services.conversion import point96_to_24h


# ---------------------------------------------------------------------------
# CustomerDailyConsumption
# ---------------------------------------------------------------------------

def _hours_from_model(obj: CustomerDailyConsumption) -> dict[str, Decimal]:
    return {
        f"hour_{h:02d}": getattr(obj, f"hour_{h:02d}")
        for h in range(24)
    }


def _apply_hours(obj, hours: Optional[dict[str, Decimal]]) -> None:
    if not hours:
        return
    for key, value in hours.items():
        if hasattr(obj, key):
            setattr(obj, key, value)


def _daily_out(obj: CustomerDailyConsumption) -> dict:
    """Convert ORM object to frontend-friendly dict."""
    data = {
        "id": obj.id,
        "customer_account_id": obj.customer_account_id,
        "inquiry_id": obj.inquiry_id,
        "customer_name": obj.customer_name,
        "account_number": obj.account_number,
        "data_date": obj.data_date,
        "data_month": obj.data_month,
        "hours": _hours_from_model(obj),
        "total_consumption": obj.total_consumption,
        "peak_consumption": obj.peak_consumption,
        "high_consumption": obj.high_consumption,
        "normal_consumption": obj.normal_consumption,
        "valley_consumption": obj.valley_consumption,
        "data_type": obj.data_type,
        "data_source": obj.data_source,
        "package_type": obj.package_type,
        "price_difference": obj.price_difference,
        "import_file_name": obj.import_file_name,
        "import_batch_id": obj.import_batch_id,
        "raw_data_count": obj.raw_data_count,
        "remarks": obj.remarks,
        "commission_status": obj.commission_status,
        "data_locked": obj.data_locked,
        "commission_calculated_time": obj.commission_calculated_time,
        "region": obj.region,
        "created_at": obj.created_at,
        "updated_at": obj.updated_at,
    }
    return data


class CustomerDailyConsumptionService:
    """CRUD service for CustomerDailyConsumption."""

    @staticmethod
    def list_page(
        db: Session,
        page: int = 1,
        page_size: int = 20,
        customer_account_id: Optional[int] = None,
        data_month: Optional[str] = None,
        data_date: Optional[date] = None,
    ) -> dict:
        q = db.query(CustomerDailyConsumption).filter(
            CustomerDailyConsumption.deleted_at.is_(None)
        )
        if customer_account_id is not None:
            q = q.filter(CustomerDailyConsumption.customer_account_id == customer_account_id)
        if data_month:
            q = q.filter(CustomerDailyConsumption.data_month == data_month)
        if data_date:
            q = q.filter(CustomerDailyConsumption.data_date == data_date)
        q = q.order_by(CustomerDailyConsumption.data_date.desc())
        page_result = paginate_query(db, q, page, page_size)
        page_result["list"] = [_daily_out(item) for item in page_result["list"]]
        return page_result

    @staticmethod
    def get(db: Session, record_id: int) -> Optional[dict]:
        obj = (
            db.query(CustomerDailyConsumption)
            .filter(
                CustomerDailyConsumption.id == record_id,
                CustomerDailyConsumption.deleted_at.is_(None),
            )
            .first()
        )
        return _daily_out(obj) if obj else None

    @staticmethod
    def create(db: Session, data: CustomerDailyConsumptionCreate) -> CustomerDailyConsumption:
        payload = data.model_dump(exclude={"hours"}, exclude_none=True)
        obj = CustomerDailyConsumption(**payload)
        _apply_hours(obj, data.hours)
        db.add(obj)
        db.commit()
        db.refresh(obj)
        return obj

    @staticmethod
    def update(db: Session, data: CustomerDailyConsumptionUpdate) -> Optional[CustomerDailyConsumption]:
        obj = (
            db.query(CustomerDailyConsumption)
            .filter(
                CustomerDailyConsumption.id == data.id,
                CustomerDailyConsumption.deleted_at.is_(None),
            )
            .first()
        )
        if not obj:
            return None
        for k, v in data.model_dump(exclude={"id", "hours"}, exclude_none=True).items():
            setattr(obj, k, v)
        _apply_hours(obj, data.hours)
        db.commit()
        db.refresh(obj)
        return obj

    @staticmethod
    def delete(db: Session, record_id: int) -> bool:
        obj = (
            db.query(CustomerDailyConsumption)
            .filter(
                CustomerDailyConsumption.id == record_id,
                CustomerDailyConsumption.deleted_at.is_(None),
            )
            .first()
        )
        if not obj:
            return False
        obj.deleted_at = datetime.now(timezone.utc)
        db.commit()
        return True

    @staticmethod
    def batch_create(db: Session, items: list[CustomerDailyConsumptionCreate]) -> int:
        count = 0
        for item in items:
            payload = item.model_dump(exclude={"hours"}, exclude_none=True)
            obj = CustomerDailyConsumption(**payload)
            _apply_hours(obj, item.hours)
            db.add(obj)
            count += 1
        db.commit()
        return count

    @staticmethod
    def statistics(
        db: Session,
        customer_account_id: Optional[int] = None,
        data_month: Optional[str] = None,
    ) -> dict:
        q = db.query(CustomerDailyConsumption).filter(
            CustomerDailyConsumption.deleted_at.is_(None)
        )
        if customer_account_id is not None:
            q = q.filter(CustomerDailyConsumption.customer_account_id == customer_account_id)
        if data_month:
            q = q.filter(CustomerDailyConsumption.data_month == data_month)

        total_days = q.count()
        total = q.with_entities(func.sum(CustomerDailyConsumption.total_consumption)).scalar() or Decimal("0")
        peak = q.with_entities(func.sum(CustomerDailyConsumption.peak_consumption)).scalar() or Decimal("0")
        high = q.with_entities(func.sum(CustomerDailyConsumption.high_consumption)).scalar() or Decimal("0")
        normal = q.with_entities(func.sum(CustomerDailyConsumption.normal_consumption)).scalar() or Decimal("0")
        valley = q.with_entities(func.sum(CustomerDailyConsumption.valley_consumption)).scalar() or Decimal("0")
        return {
            "total_days": total_days,
            "total_consumption": total,
            "peak_consumption": peak,
            "high_consumption": high,
            "normal_consumption": normal,
            "valley_consumption": valley,
        }


# ---------------------------------------------------------------------------
# CustomerHourlyConsumption
# ---------------------------------------------------------------------------

def _hourly_out(obj: CustomerHourlyConsumption) -> dict:
    return {
        "id": obj.id,
        "customer_account_id": obj.customer_account_id,
        "inquiry_id": obj.inquiry_id,
        "customer_name": obj.customer_name,
        "data_date": obj.data_date,
        "data_month": obj.data_month,
        "hour_index": obj.hour_index,
        "consumption": obj.consumption,
        "time_period": obj.time_period,
        "data_type": obj.data_type,
        "data_source": obj.data_source,
        "package_type": obj.package_type,
        "retail_unit_price": obj.retail_unit_price,
        "delivered_unit_price": obj.delivered_unit_price,
        "wholesale_unit_price": obj.wholesale_unit_price,
        "remarks": obj.remarks,
        "region": obj.region,
        "created_at": obj.created_at,
        "updated_at": obj.updated_at,
    }

    @staticmethod
    def export_excel(
        db: Session,
        customer_account_id: Optional[int] = None,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
    ) -> StreamingResponse:
        """Export daily consumption data as Excel."""
        from app.services.excel_utils import export_to_response

        q = db.query(CustomerDailyConsumption).filter(
            CustomerDailyConsumption.deleted_at.is_(None)
        )
        if customer_account_id:
            q = q.filter(CustomerDailyConsumption.customer_account_id == customer_account_id)
        if start_date:
            q = q.filter(CustomerDailyConsumption.data_date >= date.fromisoformat(start_date))
        if end_date:
            q = q.filter(CustomerDailyConsumption.data_date <= date.fromisoformat(end_date))

        rows = q.order_by(CustomerDailyConsumption.data_date.desc()).limit(10000).all()

        headers = [
            "ID", "客户ID", "客户名称", "代理商", "数据日期",
            "峰时段用电", "平时段用电", "谷时段用电", "尖峰用电",
            "总用电量", "加工日期",
        ]
        data = []
        for r in rows:
            data.append([
                r.id, r.customer_account_id, r.customer_name or "",
                r.agent_name or "", str(r.data_date or ""),
                float(r.peak_consumption or 0),
                float(r.normal_consumption or 0),
                float(r.valley_consumption or 0),
                float(r.high_consumption or 0),
                float(r.total_consumption or 0),
                str(r.process_date or ""),
            ])
        filename = "日用电量数据.xlsx"
        return export_to_response(headers, data, filename, "日用电量")


class CustomerHourlyConsumptionService:
    """CRUD service for CustomerHourlyConsumption."""

    @staticmethod
    def list_page(
        db: Session,
        page: int = 1,
        page_size: int = 20,
        customer_account_id: Optional[int] = None,
        data_month: Optional[str] = None,
        data_date: Optional[date] = None,
    ) -> dict:
        q = db.query(CustomerHourlyConsumption).filter(
            CustomerHourlyConsumption.deleted_at.is_(None)
        )
        if customer_account_id is not None:
            q = q.filter(CustomerHourlyConsumption.customer_account_id == customer_account_id)
        if data_month:
            q = q.filter(CustomerHourlyConsumption.data_month == data_month)
        if data_date:
            q = q.filter(CustomerHourlyConsumption.data_date == data_date)
        q = q.order_by(CustomerHourlyConsumption.data_date.desc(), CustomerHourlyConsumption.hour_index.asc())
        page_result = paginate_query(db, q, page, page_size)
        page_result["list"] = [_hourly_out(item) for item in page_result["list"]]
        return page_result

    @staticmethod
    def get(db: Session, record_id: int) -> Optional[dict]:
        obj = (
            db.query(CustomerHourlyConsumption)
            .filter(
                CustomerHourlyConsumption.id == record_id,
                CustomerHourlyConsumption.deleted_at.is_(None),
            )
            .first()
        )
        return _hourly_out(obj) if obj else None

    @staticmethod
    def create(
        db: Session, data: CustomerHourlyConsumptionCreate
    ) -> CustomerHourlyConsumption:
        obj = CustomerHourlyConsumption(**data.model_dump())
        db.add(obj)
        db.commit()
        db.refresh(obj)
        return obj

    @staticmethod
    def update(
        db: Session, data: CustomerHourlyConsumptionUpdate
    ) -> Optional[CustomerHourlyConsumption]:
        obj = (
            db.query(CustomerHourlyConsumption)
            .filter(
                CustomerHourlyConsumption.id == data.id,
                CustomerHourlyConsumption.deleted_at.is_(None),
            )
            .first()
        )
        if not obj:
            return None
        for k, v in data.model_dump(exclude={"id"}, exclude_none=True).items():
            setattr(obj, k, v)
        db.commit()
        db.refresh(obj)
        return obj

    @staticmethod
    def delete(db: Session, record_id: int) -> bool:
        obj = (
            db.query(CustomerHourlyConsumption)
            .filter(
                CustomerHourlyConsumption.id == record_id,
                CustomerHourlyConsumption.deleted_at.is_(None),
            )
            .first()
        )
        if not obj:
            return False
        obj.deleted_at = datetime.now(timezone.utc)
        db.commit()
        return True

    @staticmethod
    def export_excel(
        db: Session,
        customer_account_id: Optional[int] = None,
        data_date: Optional[str] = None,
    ) -> StreamingResponse:
        """Export hourly consumption data as Excel."""
        from app.services.excel_utils import export_to_response

        q = db.query(CustomerHourlyConsumption).filter(
            CustomerHourlyConsumption.deleted_at.is_(None)
        )
        if customer_account_id:
            q = q.filter(CustomerHourlyConsumption.customer_account_id == customer_account_id)
        if data_date:
            q = q.filter(CustomerHourlyConsumption.data_date == date.fromisoformat(data_date))

        rows = q.order_by(
            CustomerHourlyConsumption.data_date,
            CustomerHourlyConsumption.hour,
        ).limit(10000).all()

        headers = [
            "ID", "客户ID", "客户名称", "数据日期", "小时",
            "时间起", "时间止", "用电量(kWh)", "时段类型",
            "是否有效",
        ]
        data = []
        for r in rows:
            data.append([
                r.id, r.customer_account_id, r.customer_name or "",
                str(r.data_date or ""), r.hour,
                r.time_start or "", r.time_end or "",
                float(r.hourly_consumption or 0),
                r.time_period_name or "",
                "是" if r.is_valid else "否",
            ])
        filename = "小时用电量数据.xlsx"
        return export_to_response(headers, data, filename, "小时用电量")


# ---------------------------------------------------------------------------
# Point96Data
# ---------------------------------------------------------------------------

def _point96_out(obj: Point96Data) -> dict:
    """Build frontend-friendly output for Point96Data."""
    points = {}
    for name in dir(obj):
        if name.startswith("p") and len(name) == 5 and name[1:].isdigit():
            val = getattr(obj, name)
            if val is not None:
                hour = int(name[1:3])
                minute = int(name[3:5])
                label = f"{hour:02d}:{minute:02d}"
                points[label] = val
    return {
        "id": obj.id,
        "customer_account_id": obj.customer_account_id,
        "market_member_name": obj.market_member_name,
        "account_number": obj.account_number,
        "measure_point": obj.measure_point,
        "data_date": obj.data_date,
        "is_contracted": obj.is_contracted,
        "trading_unit_name": obj.trading_unit_name,
        "total_consumption": obj.total_consumption,
        "points": points,
        "batch_no": obj.batch_no,
        "processed": obj.processed,
        "convert_time": obj.convert_time,
        "region": obj.region,
        "created_at": obj.created_at,
        "updated_at": obj.updated_at,
    }


class Point96DataService:
    """CRUD + import/convert service for Point96Data."""

    @staticmethod
    def list_page(
        db: Session,
        page: int = 1,
        page_size: int = 20,
        customer_account_id: Optional[int] = None,
        data_month: Optional[str] = None,
        data_date: Optional[date] = None,
    ) -> dict:
        q = db.query(Point96Data).filter(Point96Data.deleted_at.is_(None))
        if customer_account_id is not None:
            q = q.filter(Point96Data.customer_account_id == customer_account_id)
        if data_date:
            q = q.filter(Point96Data.data_date == data_date)
        elif data_month:
            from datetime import datetime as _dt

            start = _dt.strptime(data_month, "%Y-%m").date()
            if start.month == 12:
                end = date(start.year + 1, 1, 1)
            else:
                end = date(start.year, start.month + 1, 1)
            q = q.filter(Point96Data.data_date >= start, Point96Data.data_date < end)
        q = q.order_by(Point96Data.data_date.desc())
        page_result = paginate_query(db, q, page, page_size)
        page_result["list"] = [_point96_out(item) for item in page_result["list"]]
        return page_result

    @staticmethod
    def get(db: Session, record_id: int) -> Optional[dict]:
        obj = (
            db.query(Point96Data)
            .filter(Point96Data.id == record_id, Point96Data.deleted_at.is_(None))
            .first()
        )
        return _point96_out(obj) if obj else None

    @staticmethod
    def create(db: Session, data: Point96DataCreate) -> Point96Data:
        payload = data.model_dump(exclude={"points"}, exclude_none=True)
        obj = Point96Data(**payload)
        if data.points:
            for label, value in data.points.items():
                hour, minute = label.split(":")
                attr = f"p{hour}{minute}"
                if hasattr(obj, attr):
                    setattr(obj, attr, value)
        db.add(obj)
        db.commit()
        db.refresh(obj)
        return obj

    @staticmethod
    def delete(db: Session, record_id: int) -> bool:
        obj = (
            db.query(Point96Data)
            .filter(Point96Data.id == record_id, Point96Data.deleted_at.is_(None))
            .first()
        )
        if not obj:
            return False
        obj.deleted_at = datetime.now(timezone.utc)
        db.commit()
        return True

    @staticmethod
    def import_from_file(db: Session, file: UploadFile) -> dict:
        """Parse uploaded 96-point data file and create Point96Data rows.

        Expected Excel format (header row optional):
          customer_account_id | data_date | customer_name | point_01 ... point_96
        """
        import openpyxl
        from io import BytesIO

        if not file.filename or not file.filename.lower().endswith(".xlsx"):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Only .xlsx files are supported",
            )
        content = file.file.read()
        workbook = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

        created = 0
        errors = []
        for row_idx, row in enumerate(rows, 1):
            if row_idx == 1:
                # Check if first row is a header (non-numeric first cell)
                if row and row[0] and not isinstance(row[0], (int, float)):
                    continue
            if not row or row[0] is None:
                continue

            try:
                customer_account_id = int(row[0])
                data_date = row[1]
                if isinstance(data_date, str):
                    data_date = date.fromisoformat(data_date)
                elif hasattr(data_date, 'date'):
                    data_date = data_date.date()
                elif isinstance(data_date, datetime):
                    data_date = data_date.date()
                else:
                    data_date = date.today()

                # Parse 96 point values (columns 3 to 98)
                point_values = [0.0] * 96
                customer_name = str(row[2]) if len(row) > 2 and row[2] else None
                for i in range(96):
                    col_idx = 3 + i
                    if col_idx < len(row) and row[col_idx] is not None:
                        point_values[i] = float(row[col_idx])

                obj = Point96Data(
                    customer_account_id=customer_account_id,
                    customer_name=customer_name,
                    data_date=data_date,
                    point_01=point_values[0], point_02=point_values[1],
                    point_03=point_values[2], point_04=point_values[3],
                    point_05=point_values[4], point_06=point_values[5],
                    point_07=point_values[6], point_08=point_values[7],
                    point_09=point_values[8], point_10=point_values[9],
                    point_11=point_values[10], point_12=point_values[11],
                    point_13=point_values[12], point_14=point_values[13],
                    point_15=point_values[14], point_16=point_values[15],
                    point_17=point_values[16], point_18=point_values[17],
                    point_19=point_values[18], point_20=point_values[19],
                    point_21=point_values[20], point_22=point_values[21],
                    point_23=point_values[22], point_24=point_values[23],
                    point_25=point_values[24], point_26=point_values[25],
                    point_27=point_values[26], point_28=point_values[27],
                    point_29=point_values[28], point_30=point_values[29],
                    point_31=point_values[30], point_32=point_values[31],
                    point_33=point_values[32], point_34=point_values[33],
                    point_35=point_values[34], point_36=point_values[35],
                    point_37=point_values[36], point_38=point_values[37],
                    point_39=point_values[38], point_40=point_values[39],
                    point_41=point_values[40], point_42=point_values[41],
                    point_43=point_values[42], point_44=point_values[43],
                    point_45=point_values[44], point_46=point_values[45],
                    point_47=point_values[46], point_48=point_values[47],
                    point_49=point_values[48], point_50=point_values[49],
                    point_51=point_values[50], point_52=point_values[51],
                    point_53=point_values[52], point_54=point_values[53],
                    point_55=point_values[54], point_56=point_values[55],
                    point_57=point_values[56], point_58=point_values[57],
                    point_59=point_values[58], point_60=point_values[59],
                    point_61=point_values[60], point_62=point_values[61],
                    point_63=point_values[62], point_64=point_values[63],
                    point_65=point_values[64], point_66=point_values[65],
                    point_67=point_values[66], point_68=point_values[67],
                    point_69=point_values[68], point_70=point_values[69],
                    point_71=point_values[70], point_72=point_values[71],
                    point_73=point_values[72], point_74=point_values[73],
                    point_75=point_values[74], point_76=point_values[75],
                    point_77=point_values[76], point_78=point_values[77],
                    point_79=point_values[78], point_80=point_values[79],
                    point_81=point_values[80], point_82=point_values[81],
                    point_83=point_values[82], point_84=point_values[83],
                    point_85=point_values[84], point_86=point_values[85],
                    point_87=point_values[86], point_88=point_values[87],
                    point_89=point_values[88], point_90=point_values[89],
                    point_91=point_values[90], point_92=point_values[91],
                    point_93=point_values[92], point_94=point_values[93],
                    point_95=point_values[94], point_96=point_values[95],
                )
                db.add(obj)
                created += 1
            except Exception as e:
                errors.append({"row": row_idx, "error": str(e)})

        if created > 0:
            db.commit()
        return {"created": created, "errors": errors, "filename": file.filename}

    @staticmethod
    def convert_to_daily(db: Session, record_id: int) -> dict:
        """Convert a Point96Data record into 24h daily consumption dict."""
        obj = get_or_404(db, Point96Data, record_id)
        hours = point96_to_24h(obj)
        obj.processed = 1
        obj.convert_time = datetime.now(timezone.utc).date()
        db.commit()
        return {
            "point96_id": obj.id,
            "customer_account_id": obj.customer_account_id,
            "data_date": obj.data_date,
            "hours": hours,
        }


# ===========================================================================
# Task #46 — Missing hourly consumption & 96-point endpoints
# ===========================================================================

# ─── Additional CustomerHourlyConsumption methods ─────────────────────────

class HourlyConsumptionExtendedService:
    """Extended methods for CustomerHourlyConsumption (task #46)."""

    @staticmethod
    def list_all(
        db: Session,
        customer_account_id: Optional[int] = None,
        data_month: Optional[str] = None,
    ) -> list[dict]:
        """GET /hourly-consumption/list — list without pagination."""
        q = db.query(CustomerHourlyConsumption).filter(CustomerHourlyConsumption.deleted_at.is_(None))
        if customer_account_id:
            q = q.filter(CustomerHourlyConsumption.customer_account_id == customer_account_id)
        if data_month:
            q = q.filter(CustomerHourlyConsumption.data_month == data_month)
        items = q.order_by(CustomerHourlyConsumption.data_date.desc()).all()
        return [_hourly_out(item) for item in items]

    @staticmethod
    def get_month_data(
        db: Session, customer_account_id: int, data_month: str
    ) -> list[dict]:
        """GET /hourly-consumption/customer-month-data."""
        items = (
            db.query(CustomerHourlyConsumption)
            .filter(
                CustomerHourlyConsumption.customer_account_id == customer_account_id,
                CustomerHourlyConsumption.data_month == data_month,
                CustomerHourlyConsumption.deleted_at.is_(None),
            )
            .order_by(CustomerHourlyConsumption.data_date.asc())
            .all()
        )
        return [_hourly_out(item) for item in items]

    @staticmethod
    def get_daily_data(
        db: Session, customer_account_id: int, data_date: date
    ) -> Optional[dict]:
        """GET /hourly-consumption/customer-daily."""
        obj = (
            db.query(CustomerHourlyConsumption)
            .filter(
                CustomerHourlyConsumption.customer_account_id == customer_account_id,
                CustomerHourlyConsumption.data_date == data_date,
                CustomerHourlyConsumption.deleted_at.is_(None),
            )
            .first()
        )
        return _hourly_out(obj) if obj else None

    @staticmethod
    def get_statistics(
        db: Session, customer_account_id: int, data_month: str
    ) -> dict:
        """GET /hourly-consumption/statistics — monthly aggregated stats."""
        items = (
            db.query(CustomerHourlyConsumption)
            .filter(
                CustomerHourlyConsumption.customer_account_id == customer_account_id,
                CustomerHourlyConsumption.data_month == data_month,
                CustomerHourlyConsumption.deleted_at.is_(None),
            )
            .all()
        )
        if not items:
            return {"totalConsumption": Decimal("0"), "dayCount": 0, "avgDaily": Decimal("0"),
                    "peakTotal": Decimal("0"), "highTotal": Decimal("0"),
                    "normalTotal": Decimal("0"), "valleyTotal": Decimal("0")}

        total = sum((item.total_consumption or Decimal("0")) for item in items)
        peak = sum((item.peak_consumption or Decimal("0")) for item in items)
        high = sum((item.high_consumption or Decimal("0")) for item in items)
        normal = sum((item.normal_consumption or Decimal("0")) for item in items)
        valley = sum((item.valley_consumption or Decimal("0")) for item in items)
        day_count = len(items)
        avg_daily = total / Decimal(day_count) if day_count > 0 else Decimal("0")

        return {
            "totalConsumption": total.quantize(Decimal("0.01")),
            "dayCount": day_count,
            "avgDaily": avg_daily.quantize(Decimal("0.01")),
            "peakTotal": peak.quantize(Decimal("0.01")),
            "highTotal": high.quantize(Decimal("0.01")),
            "normalTotal": normal.quantize(Decimal("0.01")),
            "valleyTotal": valley.quantize(Decimal("0.01")),
        }

    @staticmethod
    def get_time_period_statistics(
        db: Session, customer_account_id: int, data_month: str
    ) -> dict:
        """GET /hourly-consumption/time-period-statistics."""
        stats = HourlyConsumptionExtendedService.get_statistics(db, customer_account_id, data_month)
        total = stats["totalConsumption"]
        if total == 0:
            return {"peak": {"amount": 0, "ratio": 0}, "high": {"amount": 0, "ratio": 0},
                    "normal": {"amount": 0, "ratio": 0}, "valley": {"amount": 0, "ratio": 0}}
        return {
            "peak": {"amount": stats["peakTotal"], "ratio": float(stats["peakTotal"] / total * 100)},
            "high": {"amount": stats["highTotal"], "ratio": float(stats["highTotal"] / total * 100)},
            "normal": {"amount": stats["normalTotal"], "ratio": float(stats["normalTotal"] / total * 100)},
            "valley": {"amount": stats["valleyTotal"], "ratio": float(stats["valleyTotal"] / total * 100)},
        }

    @staticmethod
    def get_trend(
        db: Session, customer_account_id: int, start_date: str, end_date: str
    ) -> list[dict]:
        """GET /hourly-consumption/trend — daily consumption trend."""
        items = (
            db.query(CustomerHourlyConsumption)
            .filter(
                CustomerHourlyConsumption.customer_account_id == customer_account_id,
                CustomerHourlyConsumption.data_date >= start_date,
                CustomerHourlyConsumption.data_date <= end_date,
                CustomerHourlyConsumption.deleted_at.is_(None),
            )
            .order_by(CustomerHourlyConsumption.data_date.asc())
            .all()
        )
        return [
            {
                "date": str(item.data_date),
                "totalConsumption": float(item.total_consumption or 0),
                "peakConsumption": float(item.peak_consumption or 0),
                "valleyConsumption": float(item.valley_consumption or 0),
            }
            for item in items
        ]

    @staticmethod
    def check_completeness(
        db: Session, customer_account_id: int, data_month: str
    ) -> dict:
        """GET /hourly-consumption/check-completeness."""
        year, month = map(int, data_month.split("-"))
        if month == 12:
            next_month = date(year + 1, 1, 1)
        else:
            next_month = date(year, month + 1, 1)
        first_day = date(year, month, 1)
        expected_days = (next_month - first_day).days

        items = (
            db.query(CustomerHourlyConsumption.data_date)
            .filter(
                CustomerHourlyConsumption.customer_account_id == customer_account_id,
                CustomerHourlyConsumption.data_month == data_month,
                CustomerHourlyConsumption.deleted_at.is_(None),
            )
            .distinct()
            .all()
        )
        found_days = len(items)
        existing_dates = {item.data_date for item in items}

        all_dates = []
        current = first_day
        while current < next_month:
            all_dates.append({
                "date": str(current),
                "hasData": current in existing_dates,
            })
            from datetime import timedelta
            current += timedelta(days=1)

        return {
            "expectedDays": expected_days,
            "foundDays": found_days,
            "isComplete": found_days >= expected_days,
            "missingDays": expected_days - found_days,
            "dates": all_dates,
        }

    @staticmethod
    def count_records(
        db: Session,
        customer_account_id: Optional[int] = None,
        data_month: Optional[str] = None,
    ) -> int:
        """GET /hourly-consumption/count-records."""
        q = db.query(func.count(CustomerHourlyConsumption.id)).filter(
            CustomerHourlyConsumption.deleted_at.is_(None)
        )
        if customer_account_id:
            q = q.filter(CustomerHourlyConsumption.customer_account_id == customer_account_id)
        if data_month:
            q = q.filter(CustomerHourlyConsumption.data_month == data_month)
        return q.scalar() or 0

    @staticmethod
    def get_daily_summary(
        db: Session, customer_account_id: int, data_month: str
    ) -> list[dict]:
        """GET /hourly-consumption/daily-summary."""
        items = (
            db.query(CustomerHourlyConsumption)
            .filter(
                CustomerHourlyConsumption.customer_account_id == customer_account_id,
                CustomerHourlyConsumption.data_month == data_month,
                CustomerHourlyConsumption.deleted_at.is_(None),
            )
            .order_by(CustomerHourlyConsumption.data_date.asc())
            .all()
        )
        return [
            {
                "date": str(item.data_date),
                "total": float(item.total_consumption or 0),
                "peak": float(item.peak_consumption or 0),
                "high": float(item.high_consumption or 0),
                "normal": float(item.normal_consumption or 0),
                "valley": float(item.valley_consumption or 0),
            }
            for item in items
        ]

    @staticmethod
    def submit_24hour_data(
        db: Session,
        customer_account_id: int,
        data_date: date,
        hours: list[Optional[Decimal]],
        customer_name: Optional[str] = None,
    ) -> dict:
        """POST /hourly-consumption/submit-24hour-data."""
        data_month = data_date.strftime("%Y-%m")
        existing = (
            db.query(CustomerHourlyConsumption)
            .filter(
                CustomerHourlyConsumption.customer_account_id == customer_account_id,
                CustomerHourlyConsumption.data_date == data_date,
                CustomerHourlyConsumption.deleted_at.is_(None),
            )
            .first()
        )
        if existing:
            obj = existing
        else:
            obj = CustomerHourlyConsumption(
                customer_account_id=customer_account_id,
                customer_name=customer_name,
                data_date=data_date,
                data_month=data_month,
            )
            db.add(obj)

        for h in range(min(24, len(hours))):
            setattr(obj, f"hour_{h:02d}", hours[h])

        # Calculate totals
        vals = [hours[h] for h in range(min(24, len(hours))) if hours[h] is not None]
        obj.total_consumption = sum(vals, Decimal("0")) if vals else Decimal("0")
        db.commit()
        db.refresh(obj)
        return _hourly_out(obj)

    @staticmethod
    def split_from_time_of_use(
        db: Session,
        customer_account_id: int,
        data_date: date,
        peak: Decimal,
        high: Decimal,
        normal: Decimal,
        valley: Decimal,
    ) -> dict:
        """POST /hourly-consumption/split-from-time-of-use.

        Distribute peak/high/normal/valley totals across 24h using time-period coefficients.
        """
        data_month = data_date.strftime("%Y-%m")
        month = data_date.month
        from app.services.pricing_engine import calculate_time_period, get_price_coefficient

        total = peak + high + normal + valley
        if total == 0:
            return {"error": "Total consumption is zero"}

        # Compute coefficient-weighted denominator
        hour_periods = {h: calculate_time_period(h, month) for h in range(24)}
        hour_coeffs = {h: get_price_coefficient(hour_periods[h]) for h in range(24)}

        # Assign hours to their period group
        peak_hours = [h for h in range(24) if hour_periods[h] == 1]
        high_hours = [h for h in range(24) if hour_periods[h] == 2]
        normal_hours = [h for h in range(24) if hour_periods[h] == 3]
        valley_hours = [h for h in range(24) if hour_periods[h] == 4]

        hours_out: dict[str, Decimal] = {}
        for period_name, period_hours, period_total in [
            ("peak", peak_hours, peak),
            ("high", high_hours, high),
            ("normal", normal_hours, normal),
            ("valley", valley_hours, valley),
        ]:
            if period_hours and period_total > 0:
                # Even distribution within period group
                per_hour = period_total / Decimal(len(period_hours))
                for h in period_hours:
                    hours_out[f"hour_{h:02d}"] = per_hour.quantize(Decimal("0.0001"))
            else:
                for h in period_hours:
                    hours_out[f"hour_{h:02d}"] = Decimal("0")

        return {
            "customer_account_id": customer_account_id,
            "data_date": str(data_date),
            "hours": hours_out,
            "total": total.quantize(Decimal("0.01")),
        }

    @staticmethod
    def delete_monthly_data(
        db: Session, customer_account_id: int, data_month: str
    ) -> int:
        """DELETE /hourly-consumption/delete-monthly-data."""
        count = (
            db.query(CustomerHourlyConsumption)
            .filter(
                CustomerHourlyConsumption.customer_account_id == customer_account_id,
                CustomerHourlyConsumption.data_month == data_month,
                CustomerHourlyConsumption.deleted_at.is_(None),
            )
            .update({CustomerHourlyConsumption.deleted_at: datetime.now(timezone.utc)}, synchronize_session=False)
        )
        db.commit()
        return count

    @staticmethod
    def delete_daily_data(
        db: Session, customer_account_id: int, data_date: date
    ) -> int:
        """DELETE /hourly-consumption/delete-daily-data."""
        count = (
            db.query(CustomerHourlyConsumption)
            .filter(
                CustomerHourlyConsumption.customer_account_id == customer_account_id,
                CustomerHourlyConsumption.data_date == data_date,
                CustomerHourlyConsumption.deleted_at.is_(None),
            )
            .update({CustomerHourlyConsumption.deleted_at: datetime.now(timezone.utc)}, synchronize_session=False)
        )
        db.commit()
        return count

    @staticmethod
    def batch_import(
        db: Session, records: list[dict]
    ) -> int:
        """POST /hourly-consumption/batch-import — batch import 24h records."""
        count = 0
        for rec in records:
            customer_account_id = rec.get("customer_account_id")
            data_date = rec.get("data_date")
            hours = rec.get("hours", [])
            if not customer_account_id or not data_date:
                continue
            HourlyConsumptionExtendedService.submit_24hour_data(
                db, customer_account_id, data_date, hours,
                rec.get("customer_name"),
            )
            count += 1
        return count

    @staticmethod
    def convert_inquiry_to_customer(
        db: Session, inquiry_id: int, customer_account_id: int
    ) -> dict:
        """POST /hourly-consumption/convert-inquiry-to-customer.
        Copy consumption data from inquiry-linked records to customer account.
        """
        from app.models.inquiry import Inquiry

        inquiry = (
            db.query(Inquiry)
            .filter(Inquiry.id == inquiry_id, Inquiry.deleted_at.is_(None))
            .first()
        )
        if not inquiry:
            raise HTTPException(status_code=404, detail="询价单不存在")

        # Get inquiry consumption data
        import json
        cons_data = None
        if inquiry.consumption_data_json:
            try:
                cons_data = json.loads(inquiry.consumption_data_json)
            except (json.JSONDecodeError, TypeError):
                pass

        if not cons_data or "hours" not in cons_data:
            raise HTTPException(status_code=400, detail="询价单无用电量数据")

        hours = cons_data["hours"]
        data_month = inquiry.usage_month or datetime.now(timezone.utc).strftime("%Y-%m")

        # Find or create hourly records for this customer
        existing = (
            db.query(CustomerHourlyConsumption)
            .filter(
                CustomerHourlyConsumption.customer_account_id == customer_account_id,
                CustomerHourlyConsumption.inquiry_id == inquiry_id,
                CustomerHourlyConsumption.deleted_at.is_(None),
            )
            .all()
        )

        # If no existing records, create from inquiry data
        if not existing:
            obj = CustomerHourlyConsumption(
                customer_account_id=customer_account_id,
                inquiry_id=inquiry_id,
                customer_name=inquiry.customer_name,
                data_month=data_month,
            )
            for h in range(min(24, len(hours))):
                val = hours[h] if hours[h] is not None else Decimal("0")
                setattr(obj, f"hour_{h:02d}", Decimal(str(val)))
            obj.total_consumption = sum(
                (Decimal(str(hours[h])) if h < len(hours) and hours[h] is not None else Decimal("0"))
                for h in range(24)
            )
            db.add(obj)
            db.commit()
            db.refresh(obj)
            return {"created": True, "id": obj.id, "customer_account_id": customer_account_id}

        return {"created": False, "existing_count": len(existing), "customer_account_id": customer_account_id}


# ─── Additional Point96 methods ───────────────────────────────────────────

class Point96ExtendedService:
    """Extended methods for Point96Data (task #46)."""

    @staticmethod
    def get_by_customer_date(
        db: Session, customer_account_id: int, data_date: date
    ) -> Optional[dict]:
        """GET /point96/get-by-customer-date."""
        obj = (
            db.query(Point96Data)
            .filter(
                Point96Data.customer_account_id == customer_account_id,
                Point96Data.data_date == data_date,
                Point96Data.deleted_at.is_(None),
            )
            .first()
        )
        return _point96_out(obj) if obj else None

    @staticmethod
    def list_by_customer(
        db: Session, customer_account_id: int
    ) -> list[dict]:
        """GET /point96/list-by-customer."""
        items = (
            db.query(Point96Data)
            .filter(
                Point96Data.customer_account_id == customer_account_id,
                Point96Data.deleted_at.is_(None),
            )
            .order_by(Point96Data.data_date.desc())
            .all()
        )
        return [_point96_out(item) for item in items]

    @staticmethod
    def delete_by_customer_date(
        db: Session, customer_account_id: int, data_date: date
    ) -> int:
        """DELETE /point96/delete-by-customer-date."""
        count = (
            db.query(Point96Data)
            .filter(
                Point96Data.customer_account_id == customer_account_id,
                Point96Data.data_date == data_date,
                Point96Data.deleted_at.is_(None),
            )
            .update({Point96Data.deleted_at: datetime.now(timezone.utc)}, synchronize_session=False)
        )
        db.commit()
        return count


# ─── Customer Savings (task #48) ──────────────────────────────────────────

class CustomerSavingsService:
    """Customer savings analysis module (task #48 — new module)."""

    @staticmethod
    def preview_savings(
        db: Session,
        customer_account_id: int,
        data_month: str,
    ) -> dict:
        """GET /customer-savings/preview.

        Compare customer total electricity fee vs grid total fee.
        Savings = grid_total - customer_total.
        """
        # Get customer's hourly consumption
        items = (
            db.query(CustomerHourlyConsumption)
            .filter(
                CustomerHourlyConsumption.customer_account_id == customer_account_id,
                CustomerHourlyConsumption.data_month == data_month,
                CustomerHourlyConsumption.deleted_at.is_(None),
            )
            .all()
        )
        if not items:
            return {
                "customer_account_id": customer_account_id,
                "data_month": data_month,
                "totalSavings": Decimal("0"),
                "totalGridFee": Decimal("0"),
                "totalCustomerFee": Decimal("0"),
                "savingsRate": Decimal("0"),
                "dailyDetails": [],
            }

        # Get grid prices for the month
        from app.models.price import GridPrice
        from app.services.pricing_engine import calculate_time_period, get_price_coefficient

        grid_rows = (
            db.query(GridPrice)
            .filter(GridPrice.year_month == data_month, GridPrice.deleted_at.is_(None))
            .all()
        )
        period_prices: dict[int, Decimal] = {}
        for row in grid_rows:
            if row.time_period is not None and row.price is not None:
                period_prices[row.time_period] = row.price

        year, month = map(int, data_month.split("-"))

        # Get customer's price difference and package type
        from app.models.customer_account import CustomerAccount
        customer = db.query(CustomerAccount).filter(
            CustomerAccount.id == customer_account_id,
            CustomerAccount.deleted_at.is_(None),
        ).first()

        price_diff = customer.price_difference if customer and customer.price_difference else Decimal("0")
        package_type = customer.package_type if customer and customer.package_type else 2

        daily_details: list[dict] = []
        total_grid_fee = Decimal("0")
        total_customer_fee = Decimal("0")
        total_consumption = Decimal("0")

        from app.models.price import BasePrice
        base_prices_rows = (
            db.query(BasePrice)
            .filter(
                BasePrice.price_date >= date(year, month, 1),
                BasePrice.price_date < (date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)),
                BasePrice.deleted_at.is_(None),
            )
            .all()
        )
        base_hourly: dict[int, Decimal] = {}
        for row in base_prices_rows:
            if row.hour_index is not None and row.price is not None:
                base_hourly[row.hour_index] = row.price

        for item in items:
            day_grid = Decimal("0")
            day_customer = Decimal("0")
            day_consumption = Decimal("0")

            for h in range(24):
                consumption = getattr(item, f"hour_{h:02d}", None) or Decimal("0")
                period = calculate_time_period(h, month)
                grid_price = period_prices.get(period) or period_prices.get(3) or Decimal("0")

                if package_type == 1:
                    # Flat-rate
                    normal_price = (sum(base_hourly.values(), Decimal("0")) / Decimal(len(base_hourly))) if base_hourly else Decimal("0")
                    coeff = get_price_coefficient(period)
                    customer_price = normal_price * coeff + price_diff
                else:
                    # Timed
                    customer_price = base_hourly.get(h, Decimal("0")) + price_diff

                day_grid += grid_price * consumption
                day_customer += customer_price * consumption
                day_consumption += consumption

            day_savings = day_grid - day_customer
            total_grid_fee += day_grid
            total_customer_fee += day_customer
            total_consumption += day_consumption

            daily_details.append({
                "date": str(item.data_date),
                "consumption": float(day_consumption),
                "gridFee": float(day_grid),
                "customerFee": float(day_customer),
                "savings": float(day_savings),
            })

        total_savings = total_grid_fee - total_customer_fee
        savings_rate = (total_savings / total_grid_fee * 100) if total_grid_fee > 0 else Decimal("0")

        return {
            "customer_account_id": customer_account_id,
            "data_month": data_month,
            "totalSavings": total_savings.quantize(Decimal("0.01")),
            "totalGridFee": total_grid_fee.quantize(Decimal("0.01")),
            "totalCustomerFee": total_customer_fee.quantize(Decimal("0.01")),
            "totalConsumption": total_consumption.quantize(Decimal("0.01")),
            "savingsRate": savings_rate.quantize(Decimal("0.01")),
            "dayCount": len(daily_details),
            "dailyDetails": daily_details,
        }

    @staticmethod
    def export_excel(
        db: Session, customer_account_id: int, data_month: str
    ) -> bytes:
        """GET /customer-savings/export-excel."""
        from io import BytesIO
        from openpyxl import Workbook

        data = CustomerSavingsService.preview_savings(db, customer_account_id, data_month)

        wb = Workbook()
        ws = wb.active
        ws.title = "客户节约详情"

        ws.append(["客户账户ID", customer_account_id])
        ws.append(["用电月份", data_month])
        ws.append(["总节约金额(元)", float(data["totalSavings"])])
        ws.append(["国网总费用(元)", float(data["totalGridFee"])])
        ws.append(["客户总费用(元)", float(data["totalCustomerFee"])])
        ws.append(["总用电量(kWh)", float(data["totalConsumption"])])
        ws.append(["节约率(%)", float(data["savingsRate"])])
        ws.append([])

        ws.append(["日期", "用电量(kWh)", "国网费用(元)", "客户费用(元)", "节约金额(元)"])
        for detail in data["dailyDetails"]:
            ws.append([
                detail["date"],
                detail["consumption"],
                detail["gridFee"],
                detail["customerFee"],
                detail["savings"],
            ])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
